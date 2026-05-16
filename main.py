from fastapi import FastAPI, Depends, HTTPException, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, RedirectResponse
from pydantic import BaseModel
from typing import Optional
import firebase_admin
from firebase_admin import credentials, firestore
import hashlib, os, json, uuid, unicodedata, requests
from datetime import datetime, timedelta, date
from jose import jwt, JWTError
from io import BytesIO
import openpyxl
from apscheduler.schedulers.background import BackgroundScheduler

# ── CONFIG ────────────────────────────────────────────────────
JWT_SECRET      = os.environ.get("JWT_SECRET", "cambia-esto-en-render")
JWT_ALGORITHM   = "HS256"
JWT_HOURS       = 12
RETENTION_YEARS = 4
AZURE_CLIENT_ID     = os.environ.get("AZURE_CLIENT_ID", "")
AZURE_CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET", "")
APP_URL         = "https://drt-fichaje.onrender.com"
REDIRECT_URI    = f"{APP_URL}/api/auth/onedrive/callback"
GRAPH_SCOPE     = "Files.ReadWrite offline_access"

app = FastAPI(title="DRT Fichaje API", docs_url=None, redoc_url=None)

# ── FIREBASE ──────────────────────────────────────────────────
_creds_json = os.environ.get("FIREBASE_CREDENTIALS")
if _creds_json:
    _cred = credentials.Certificate(json.loads(_creds_json))
    firebase_admin.initialize_app(_cred)
    db = firestore.client()
else:
    db = None

# ── HELPERS GENERALES ─────────────────────────────────────────
def norm(s): return unicodedata.normalize("NFD", s.strip().lower()).encode("ascii","ignore").decode()
def uid(): return uuid.uuid4().hex[:20]
def hash_pwd(p): return hashlib.sha256(f"DRT:{p}".encode()).hexdigest()
def get_ip(r): fwd=r.headers.get("X-Forwarded-For"); return fwd.split(",")[0].strip() if fwd else (r.client.host if r.client else "N/D")
def check_db():
    if not db: raise HTTPException(503,"Base de datos no disponible")
def audit_log(event, user_id, details=""):
    db.collection("audit").add({"type":event,"userId":user_id,"ts":datetime.utcnow().isoformat(),"details":details})

# ── AUTH JWT ──────────────────────────────────────────────────
bearer = HTTPBearer()
def make_token(uid, role): return jwt.encode({"sub":uid,"role":role,"exp":datetime.utcnow()+timedelta(hours=JWT_HOURS)},JWT_SECRET,algorithm=JWT_ALGORITHM)
def verify_token(creds: HTTPAuthorizationCredentials = Depends(bearer)):
    try: return jwt.decode(creds.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except JWTError: raise HTTPException(401,"Token inválido o expirado")
def require_admin(payload=Depends(verify_token)):
    if payload.get("role")!="admin": raise HTTPException(403,"Acceso denegado")
    return payload

# ── ONEDRIVE OAUTH ────────────────────────────────────────────
def get_auth_url(state):
    from urllib.parse import urlencode
    p={"client_id":AZURE_CLIENT_ID,"response_type":"code","redirect_uri":REDIRECT_URI,"scope":GRAPH_SCOPE,"state":state,"response_mode":"query"}
    return f"https://login.microsoftonline.com/common/oauth2/v2.0/authorize?{urlencode(p)}"

def exchange_code(code):
    r=requests.post("https://login.microsoftonline.com/common/oauth2/v2.0/token",data={"client_id":AZURE_CLIENT_ID,"client_secret":AZURE_CLIENT_SECRET,"code":code,"redirect_uri":REDIRECT_URI,"grant_type":"authorization_code"})
    return r.json()

def do_refresh(refresh_tok):
    r=requests.post("https://login.microsoftonline.com/common/oauth2/v2.0/token",data={"client_id":AZURE_CLIENT_ID,"client_secret":AZURE_CLIENT_SECRET,"refresh_token":refresh_tok,"grant_type":"refresh_token","scope":GRAPH_SCOPE})
    return r.json()

def get_valid_token(user_id):
    doc=db.collection("users").document(user_id).get()
    if not doc.exists: raise HTTPException(404,"Usuario no encontrado")
    u=doc.to_dict()
    at=u.get("onedrive_access_token"); rt=u.get("onedrive_refresh_token"); exp=u.get("onedrive_token_expiry","")
    if not at or not rt: raise HTTPException(400,"OneDrive no conectado")
    if exp:
        try:
            if datetime.utcnow()>=datetime.fromisoformat(exp)-timedelta(minutes=5):
                nt=do_refresh(rt)
                if "access_token" in nt:
                    ei=nt.get("expires_in",3600)
                    db.collection("users").document(user_id).update({"onedrive_access_token":nt["access_token"],"onedrive_refresh_token":nt.get("refresh_token",rt),"onedrive_token_expiry":(datetime.utcnow()+timedelta(seconds=ei)).isoformat()})
                    return nt["access_token"]
        except: pass
    return at

def upload_to_onedrive(access_token, folder, filename, content):
    url=f"https://graph.microsoft.com/v1.0/me/drive/root:/DRT-Fichaje/{folder}/{filename}:/content"
    r=requests.put(url,headers={"Authorization":f"Bearer {access_token}","Content-Type":"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},data=content)
    return r.status_code in [200,201]

# ── EXCEL ─────────────────────────────────────────────────────
def fmt_ts(ts):
    try: dt=datetime.fromisoformat(ts); return dt.strftime("%d/%m/%Y"),dt.strftime("%H:%M:%S")
    except: return ts,""

def gen_excel(users_data, fichajes_data, audit_data):
    wb=openpyxl.Workbook()
    umap={u["id"]:u.get("fullname","") for u in users_data}
    # Fichajes
    ws1=wb.active; ws1.title="Fichajes"
    ws1.append(["Empleado","Tipo","Fecha","Hora","IP","Latitud","Longitud"])
    for f in fichajes_data:
        fecha,hora=fmt_ts(f.get("ts",""))
        ws1.append([umap.get(f.get("uid",""),""),f.get("tipo",""),fecha,hora,f.get("ip",""),f.get("lat",""),f.get("lon","")])
    # Empleados
    ws2=wb.create_sheet("Empleados")
    ws2.append(["Nombre","Apellido","Email","Teléfono","Departamento","Rol","Alta"])
    for u in users_data:
        alta,_=fmt_ts(u.get("createdAt",""))
        ws2.append([u.get("nombre",""),u.get("apellido",""),u.get("email",""),u.get("phone",""),u.get("dept",""),u.get("role",""),alta])
    # Auditoría
    ws3=wb.create_sheet("Auditoría")
    ws3.append(["Evento","Fecha/Hora","Detalle"])
    for a in audit_data:
        fecha,hora=fmt_ts(a.get("ts",""))
        ws3.append([a.get("type",""),f"{fecha} {hora}",a.get("details","")])
    buf=BytesIO(); wb.save(buf); buf.seek(0); return buf.read()

def get_all_data():
    users=[{"id":d.id,**{k:v for k,v in d.to_dict().items() if k not in ["hash","onedrive_access_token","onedrive_refresh_token","onedrive_token_expiry"]}} for d in db.collection("users").stream()]
    fichajes=[{"id":d.id,**d.to_dict()} for d in db.collection("fichajes").stream()]
    audit=[{"id":d.id,**d.to_dict()} for d in db.collection("audit").order_by("ts",direction=firestore.Query.DESCENDING).limit(1000).stream()]
    return users,fichajes,audit

# ── BACKUP AUTOMÁTICO ─────────────────────────────────────────
def run_backup_all(tipo):
    if not db: return
    try:
        admins=list(db.collection("users").where("role","==","admin").where("onedrive_connected","==",True).stream())
        today=date.today()
        for adoc in admins:
            admin={"id":adoc.id,**adoc.to_dict()}
            try:
                at=get_valid_token(admin["id"])
                all_users,all_fichajes,all_audit=get_all_data()
                if tipo=="daily":
                    ayer=(today-timedelta(days=1)).isoformat()
                    fich=[f for f in all_fichajes if f.get("ts","").startswith(ayer)]
                    folder,filename="Diario",f"{ayer}.xlsx"
                elif tipo=="monthly":
                    mes=today.strftime("%Y-%m")
                    fich=[f for f in all_fichajes if f.get("ts","").startswith(mes)]
                    folder,filename="Mensual",f"{mes}.xlsx"
                else:
                    year=str(today.year)
                    fich=[f for f in all_fichajes if f.get("ts","").startswith(year)]
                    folder,filename="Anual",f"{year}.xlsx"
                content=gen_excel(all_users,fich,all_audit)
                if upload_to_onedrive(at,folder,filename,content):
                    audit_log("BACKUP_AUTO",admin["id"],f"Backup automático {tipo}: {folder}/{filename}")
            except: pass
    except: pass

scheduler=BackgroundScheduler(timezone="Europe/Madrid")
scheduler.add_job(lambda:run_backup_all("daily"),"cron",hour=23,minute=0)
scheduler.add_job(lambda:run_backup_all("monthly"),"cron",day=1,hour=0,minute=30)
scheduler.add_job(lambda:run_backup_all("annual"),"cron",month=1,day=1,hour=1,minute=0)
scheduler.start()

# ── MODELOS ───────────────────────────────────────────────────
class LoginReq(BaseModel): nombre:str; password:str
class ChangePwdReq(BaseModel): new_pwd:str
class FicharReq(BaseModel): lat:Optional[float]=None; lon:Optional[float]=None; acc:Optional[float]=None
class SetupReq(BaseModel): nombre:str; apellido:str; email:Optional[str]=""; dept:Optional[str]="Dirección"; pwd:str
class CreateUserReq(BaseModel): nombre:str; apellido:str; email:Optional[str]=""; phone:Optional[str]=""; dept:Optional[str]=""; role:str="employee"; temp_pwd:str
class ResetPwdReq(BaseModel): user_id:str; new_pwd:str
class UpdateUserReq(BaseModel): active:Optional[bool]=None; dept:Optional[str]=None; email:Optional[str]=None; phone:Optional[str]=None
class UpdateProfileReq(BaseModel): nombre:Optional[str]=None; apellido:Optional[str]=None; email:Optional[str]=None; phone:Optional[str]=None; dept:Optional[str]=None; backup_time:Optional[str]=None

# ── ENDPOINTS PÚBLICOS ────────────────────────────────────────
@app.get("/api/status")
async def status():
    if not db: return {"configured":False}
    has=bool(list(db.collection("users").limit(1).stream()))
    return {"configured":has}

@app.post("/api/setup")
async def setup(data:SetupReq):
    check_db()
    if list(db.collection("users").limit(1).stream()): raise HTTPException(400,"Ya configurado")
    if len(data.pwd)<6: raise HTTPException(400,"Contraseña mínimo 6 caracteres")
    fullname=f"{data.nombre.strip()} {data.apellido.strip()}"
    user_id=uid()
    user={"nombre":data.nombre.strip(),"apellido":data.apellido.strip(),"fullname":fullname,"loginKey":norm(fullname),"email":data.email or "","phone":"","dept":data.dept or "Dirección","role":"admin","hash":hash_pwd(data.pwd),"active":True,"firstLogin":False,"failedAttempts":0,"createdAt":datetime.utcnow().isoformat(),"onedrive_connected":False,"backup_time":"23:00"}
    db.collection("users").document(user_id).set(user)
    audit_log("SETUP",user_id,"Sistema inicializado")
    return {"token":make_token(user_id,"admin"),"user":{**{k:v for k,v in user.items() if k not in ["hash","onedrive_access_token","onedrive_refresh_token"]},"id":user_id}}

@app.post("/api/auth/login")
async def login(data:LoginReq, request:Request):
    check_db()
    key=norm(data.nombre)
    docs=list(db.collection("users").where("loginKey","==",key).limit(1).stream())
    if not docs: raise HTTPException(401,"Empleado no encontrado. Escribe nombre y apellido exactos.")
    u={"id":docs[0].id,**docs[0].to_dict()}
    if not u.get("active",True): raise HTTPException(403,"Cuenta desactivada.")
    locked=u.get("lockedUntil")
    if locked and datetime.fromisoformat(locked)>datetime.utcnow():
        mins=max(1,int((datetime.fromisoformat(locked)-datetime.utcnow()).seconds/60)+1)
        raise HTTPException(429,f"Cuenta bloqueada {mins} min.")
    if hash_pwd(data.password)!=u.get("hash"):
        failed=u.get("failedAttempts",0)+1
        upd={"failedAttempts":failed}
        if failed>=3: upd["lockedUntil"]=(datetime.utcnow()+timedelta(minutes=5)).isoformat()
        db.collection("users").document(u["id"]).update(upd)
        raise HTTPException(401,f"Contraseña incorrecta. Intento {failed}/3.")
    db.collection("users").document(u["id"]).update({"failedAttempts":0,"lockedUntil":None})
    audit_log("LOGIN",u["id"],f"Login: {u['fullname']}")
    return {"token":make_token(u["id"],u.get("role","employee")),"user":{k:v for k,v in u.items() if k not in ["hash","onedrive_access_token","onedrive_refresh_token","onedrive_token_expiry"]}}

# ── ENDPOINTS EMPLEADO ────────────────────────────────────────
@app.get("/api/me")
async def get_me(payload=Depends(verify_token)):
    check_db()
    doc=db.collection("users").document(payload["sub"]).get()
    if not doc.exists: raise HTTPException(404,"No encontrado")
    u=doc.to_dict()
    return {k:v for k,v in {**u,"id":doc.id}.items() if k not in ["hash","onedrive_access_token","onedrive_refresh_token","onedrive_token_expiry"]}

@app.post("/api/auth/change-password")
async def change_password(data:ChangePwdReq, payload=Depends(verify_token)):
    check_db()
    if len(data.new_pwd)<6: raise HTTPException(400,"Mínimo 6 caracteres")
    db.collection("users").document(payload["sub"]).update({"hash":hash_pwd(data.new_pwd),"firstLogin":False})
    audit_log("FIRST_PWD",payload["sub"],"Contraseña personal creada")
    return {"ok":True}

@app.put("/api/profile")
async def update_profile(data:UpdateProfileReq, payload=Depends(verify_token)):
    check_db()
    upd={k:v for k,v in data.dict().items() if v is not None}
    if upd:
        if "nombre" in upd or "apellido" in upd:
            doc=db.collection("users").document(payload["sub"]).get().to_dict()
            nombre=upd.get("nombre",doc.get("nombre",""))
            apellido=upd.get("apellido",doc.get("apellido",""))
            fullname=f"{nombre.strip()} {apellido.strip()}"
            upd["fullname"]=fullname; upd["loginKey"]=norm(fullname)
        db.collection("users").document(payload["sub"]).update(upd)
        audit_log("UPDATE_PROFILE",payload["sub"],"Perfil actualizado")
    return {"ok":True}

@app.post("/api/fichar")
async def fichar(data:FicharReq, request:Request, payload=Depends(verify_token)):
    check_db()
    user_id=payload["sub"]; ip=get_ip(request)
    ultimos=list(db.collection("fichajes").where("uid","==",user_id).order_by("ts",direction=firestore.Query.DESCENDING).limit(1).stream())
    ultimo=ultimos[0].to_dict() if ultimos else None
    tipo="salida" if (ultimo and ultimo.get("tipo")=="entrada") else "entrada"
    if tipo=="entrada":
        hoy=datetime.utcnow().date().isoformat()
        entradas_hoy=list(db.collection("fichajes").where("uid","==",user_id).where("tipo","==","entrada").stream())
        if sum(1 for f in entradas_hoy if f.to_dict().get("ts","").startswith(hoy))>=3:
            raise HTTPException(400,"Máximo 3 jornadas registradas hoy.")
    ts=datetime.utcnow().isoformat()
    db.collection("fichajes").add({"uid":user_id,"tipo":tipo,"ts":ts,"ip":ip,"lat":data.lat,"lon":data.lon,"acc":data.acc})
    audit_log("FICHAJE",user_id,tipo)
    return {"ok":True,"tipo":tipo,"ts":ts}

@app.get("/api/fichajes")
async def mis_fichajes(payload=Depends(verify_token)):
    check_db()
    docs=db.collection("fichajes").where("uid","==",payload["sub"]).order_by("ts").stream()
    return [{"id":d.id,**d.to_dict()} for d in docs]

# ── ONEDRIVE OAUTH ────────────────────────────────────────────
@app.get("/api/auth/onedrive/connect")
async def onedrive_connect(token:str=None):
    if not token: raise HTTPException(401,"Token requerido")
    try: payload=jwt.decode(token,JWT_SECRET,algorithms=[JWT_ALGORITHM])
    except JWTError: raise HTTPException(401,"Token inválido")
    if not AZURE_CLIENT_ID: raise HTTPException(500,"Azure no configurado")
    return RedirectResponse(get_auth_url(payload["sub"]))

@app.get("/api/auth/onedrive/callback")
async def onedrive_callback(code:str=None, state:str=None, error:str=None):
    if error or not code or not state:
        return RedirectResponse(f"{APP_URL}?onedrive=error")
    tokens=exchange_code(code)
    if "access_token" not in tokens:
        return RedirectResponse(f"{APP_URL}?onedrive=error")
    ei=tokens.get("expires_in",3600)
    db.collection("users").document(state).update({"onedrive_access_token":tokens["access_token"],"onedrive_refresh_token":tokens.get("refresh_token",""),"onedrive_token_expiry":(datetime.utcnow()+timedelta(seconds=ei)).isoformat(),"onedrive_connected":True})
    audit_log("ONEDRIVE_CONNECT",state,"OneDrive conectado")
    return RedirectResponse(f"{APP_URL}?onedrive=success")

@app.post("/api/auth/onedrive/disconnect")
async def onedrive_disconnect(payload=Depends(verify_token)):
    check_db()
    db.collection("users").document(payload["sub"]).update({"onedrive_access_token":None,"onedrive_refresh_token":None,"onedrive_token_expiry":None,"onedrive_connected":False})
    audit_log("ONEDRIVE_DISCONNECT",payload["sub"],"OneDrive desconectado")
    return {"ok":True}

@app.get("/api/onedrive/status")
async def onedrive_status(payload=Depends(verify_token)):
    check_db()
    doc=db.collection("users").document(payload["sub"]).get()
    if not doc.exists: raise HTTPException(404)
    u=doc.to_dict()
    return {"connected":u.get("onedrive_connected",False),"backup_time":u.get("backup_time","23:00")}

# ── BACKUP ────────────────────────────────────────────────────
@app.post("/api/backup/{tipo}")
async def manual_backup(tipo:str, payload=Depends(require_admin)):
    check_db()
    if tipo not in ["daily","monthly","annual","full"]: raise HTTPException(400,"Tipo inválido")
    user_id=payload["sub"]
    try: at=get_valid_token(user_id)
    except: raise HTTPException(400,"OneDrive no conectado. Configúralo en tu perfil.")
    today=date.today()
    all_users,all_fichajes,all_audit=get_all_data()
    if tipo=="daily":
        ayer=(today-timedelta(days=1)).isoformat()
        fich=[f for f in all_fichajes if f.get("ts","").startswith(ayer)]
        folder,filename="Diario",f"{ayer}.xlsx"
    elif tipo=="monthly":
        mes=today.strftime("%Y-%m")
        fich=[f for f in all_fichajes if f.get("ts","").startswith(mes)]
        folder,filename="Mensual",f"{mes}.xlsx"
    elif tipo=="annual":
        year=str(today.year)
        fich=[f for f in all_fichajes if f.get("ts","").startswith(year)]
        folder,filename="Anual",f"{year}.xlsx"
    else:
        fich=all_fichajes
        folder,filename="Completo",f"{today.isoformat()}_completo.xlsx"
    content=gen_excel(all_users,fich,all_audit)
    ok=upload_to_onedrive(at,folder,filename,content)
    if ok:
        audit_log("BACKUP",user_id,f"Backup {tipo} → OneDrive: {folder}/{filename}")
        return {"ok":True,"file":f"DRT-Fichaje/{folder}/{filename}"}
    raise HTTPException(500,"Error al subir a OneDrive")

# ── ENDPOINTS ADMIN ───────────────────────────────────────────
@app.get("/api/admin/empleados")
async def admin_empleados(payload=Depends(require_admin)):
    check_db()
    docs=db.collection("users").stream()
    return [{"id":d.id,**{k:v for k,v in d.to_dict().items() if k not in ["hash","onedrive_access_token","onedrive_refresh_token","onedrive_token_expiry"]}} for d in docs]

@app.post("/api/admin/empleados")
async def admin_crear(data:CreateUserReq, payload=Depends(require_admin)):
    check_db()
    fullname=f"{data.nombre.strip()} {data.apellido.strip()}"
    login_key=norm(fullname)
    if list(db.collection("users").where("loginKey","==",login_key).limit(1).stream()):
        raise HTTPException(400,"Ya existe un empleado con ese nombre")
    if len(data.temp_pwd)<6: raise HTTPException(400,"Contraseña mínimo 6 caracteres")
    user_id=uid()
    user={"nombre":data.nombre.strip(),"apellido":data.apellido.strip(),"fullname":fullname,"loginKey":login_key,"email":data.email or "","phone":data.phone or "","dept":data.dept or "","role":data.role,"hash":hash_pwd(data.temp_pwd),"active":True,"firstLogin":True,"failedAttempts":0,"createdAt":datetime.utcnow().isoformat(),"createdBy":payload["sub"],"onedrive_connected":False,"backup_time":"23:00"}
    db.collection("users").document(user_id).set(user)
    audit_log("CREATE_USER",payload["sub"],f"Creado: {fullname}")
    return {"id":user_id,**{k:v for k,v in user.items() if k!="hash"}}

@app.put("/api/admin/empleados/{user_id}")
async def admin_update(user_id:str, data:UpdateUserReq, payload=Depends(require_admin)):
    check_db()
    upd={k:v for k,v in data.dict().items() if v is not None}
    if upd:
        db.collection("users").document(user_id).update(upd)
        audit_log("UPDATE_USER",payload["sub"],f"Actualizado: {user_id}")
    return {"ok":True}

@app.delete("/api/admin/empleados/{user_id}")
async def admin_delete_user(user_id:str, payload=Depends(require_admin)):
    check_db()
    if user_id==payload["sub"]: raise HTTPException(400,"No puedes eliminarte a ti mismo")
    target_doc=db.collection("users").document(user_id).get()
    if not target_doc.exists: raise HTTPException(404,"Usuario no encontrado")
    target={"id":target_doc.id,**target_doc.to_dict()}
    # Intentar backup antes de borrar
    try:
        at=get_valid_token(payload["sub"])
        all_users,all_fichajes,all_audit=get_all_data()
        user_fichajes=[f for f in all_fichajes if f.get("uid")==user_id]
        content=gen_excel([target],user_fichajes,all_audit)
        today=date.today().isoformat()
        nombre_safe=target.get("fullname","usuario").replace(" ","_")
        upload_to_onedrive(at,"Eliminados",f"{today}_{nombre_safe}.xlsx",content)
    except: pass
    # Desactivar en lugar de borrar físicamente (conservar datos)
    db.collection("users").document(user_id).update({"active":False,"deleted":True,"deletedBy":payload["sub"],"deletedAt":datetime.utcnow().isoformat()})
    audit_log("DELETE_USER",payload["sub"],f"Eliminado: {target.get('fullname','')} por {payload['sub']}")
    return {"ok":True}

@app.post("/api/admin/reset-pwd")
async def admin_reset_pwd(data:ResetPwdReq, payload=Depends(require_admin)):
    check_db()
    if len(data.new_pwd)<6: raise HTTPException(400,"Mínimo 6 caracteres")
    if not db.collection("users").document(data.user_id).get().exists: raise HTTPException(404)
    db.collection("users").document(data.user_id).update({"hash":hash_pwd(data.new_pwd),"firstLogin":True,"failedAttempts":0,"lockedUntil":None})
    audit_log("PWD_RESET",payload["sub"],f"Reset: {data.user_id}")
    return {"ok":True}

@app.get("/api/admin/fichajes")
async def admin_fichajes(payload=Depends(require_admin)):
    check_db()
    docs=db.collection("fichajes").order_by("ts",direction=firestore.Query.DESCENDING).stream()
    return [{"id":d.id,**d.to_dict()} for d in docs]

@app.delete("/api/admin/fichajes/{fichaje_id}")
async def admin_del_fichaje(fichaje_id:str, payload=Depends(require_admin)):
    check_db()
    doc=db.collection("fichajes").document(fichaje_id).get()
    if not doc.exists: raise HTTPException(404)
    ts=doc.to_dict().get("ts","")
    if ts and (datetime.utcnow()-datetime.fromisoformat(ts)).days<RETENTION_YEARS*365:
        raise HTTPException(403,f"Retención mínima {RETENTION_YEARS} años.")
    doc.reference.delete()
    audit_log("DELETE_FICHAJE",payload["sub"],fichaje_id)
    return {"ok":True}

@app.get("/api/admin/auditoria")
async def admin_auditoria(payload=Depends(require_admin)):
    check_db()
    docs=db.collection("audit").order_by("ts",direction=firestore.Query.DESCENDING).limit(500).stream()
    return [{"id":d.id,**d.to_dict()} for d in docs]

@app.get("/api/admin/export")
async def admin_export(payload=Depends(require_admin)):
    check_db()
    users,fichajes,auditoria=get_all_data()
    audit_log("EXPORT",payload["sub"],"Exportación datos")
    return {"users":users,"fichajes":fichajes,"audit":auditoria}

@app.post("/api/setup")
async def setup_check(): pass  # already defined above

# ── FRONTEND ──────────────────────────────────────────────────
import os as _os
_static=_os.path.join(_os.path.dirname(__file__),"static")
if _os.path.isdir(_static):
    app.mount("/static",StaticFiles(directory=_static),name="static")

@app.get("/{full_path:path}",include_in_schema=False)
async def serve_spa(full_path:str):
    idx=_os.path.join(_static,"index.html")
    if _os.path.isfile(idx): return FileResponse(idx)
    return {"error":"Frontend no encontrado"}
